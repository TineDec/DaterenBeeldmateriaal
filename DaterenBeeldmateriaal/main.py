import tkinter as tk
from tkinter import *
from tkinter import Listbox
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import pandas as pd

interface = tk.Tk()
interface.configure(bg="#60c1c9")
interface.geometry('1000x2000')

# kleurcode lichtblauw = #60c1c9
# kleurcode geel = #fdc20b
# kleurcode roze = #ea94aa
# kleurcode groen = #00a57e

titel = Label(interface, text="DATEREN VAN BEELDMATERIAAL")
titel.configure(bg="#60c1c9", fg="#000000", font=("Calibri", 20, "bold"))
titel.grid(row=1, column=0, columnspan=5)

# OPTIE 1: op basis van foto-techniek
# OPTIE 2: op basis van bezienswaardigheden
# doorzoek thesaurus op vaste woorden

#TO DO: SCROLLBAR TOEVOEGEN
#TO DO: ZORGEN DAT TEKST LEESBAAR IS IN TYPVELD (OP TWEE REGELS PLAATSEN?)

#wb = load_workbook(r"C:\Users\decoseti\Downloads\Gebeurtenissen_Lijst.xlsx")
#wb = load_workbook("Gebeurtenissen_Lijst.xlsx")
#ws = wb.active

#https://stackoverflow.com/questions/43922356/how-to-read-an-excel-column-into-a-list --> lijst maken van excel d.m.v. pandas?
file_name = "Gebeurtenissen_Lijst.xlsx"
xl_workbook = pd.ExcelFile(file_name)
df = xl_workbook.parse("Gebeurtenissen")
alist = df['GEBEURTENIS (PYTHON)'].tolist()

Label(
    text="Duid hier aan wat te zien is op de foto. Zoektermen moeten in het Nederlands en enkelvoud genoteerd worden.",
    bg="#60c1c9",
    fg="#000000",
    font=("Calibri", 12)
).grid(row=2, column=1, columnspan=5, pady=20)

def update(data):
    my_list_1.delete(0, END)

    for entry in data:
        my_list_1.insert(END, entry)

def check(e):
    typed = entry_1.get()

    if typed == '':
        data = alist
    else:
        data = []
        for item in alist:
            if typed.lower() in item.lower():
                data.append(item)
    update(data)

def fillout(e):
    entry_1.delete(0, END)
    entry_1.insert(0, my_list_1.get(ACTIVE))

entry_1 = Entry(interface, width=50)
entry_1.grid(row=3, column=1, padx=5, pady=5)
entry_1.bind('<KeyRelease>', check)

my_list_1: Listbox = Listbox(interface, height=20, width=50)
my_list_1.grid(row=4, column=1, padx=5, pady=5)
my_list_1.bind("<<ListboxSelect>>", fillout)

def update_2(data_2):
    my_list_2.delete(0, END)

    for item in data_2:
        my_list_2.insert(END, item)

def check_2(e):
    typed = entry_2.get()

    if typed == '':
        data_2 = alist
    else:
        data_2 = []
        for item in alist:
            if typed.lower() in item.lower():
                data_2.append(item)
    update_2(data_2)

def fillout_2(e):
    entry_2.delete(0, END)
    entry_2.insert(0, my_list_2.get(ACTIVE))


entry_2 = Entry(interface, width=50)
entry_2.grid(row=3, column=3, padx=5, pady=5)
entry_2.bind('<KeyRelease>', check_2)

my_list_2: Listbox = Listbox(interface, height=20, width=50)
my_list_2.grid(row=4, column=3, padx=5, pady=5)
my_list_2.bind("<<ListboxSelect>>", fillout_2)


def update_3(data_3):
    my_list_3.delete(0, END)

    for item in data_3:
        my_list_3.insert(END, item)

def check_3(e):
    typed = entry_3.get()

    if typed == '':
        data_3 = alist
    else:
        data_3 = []
        for item in alist:
            if typed.lower() in item.lower():
                data_3.append(item)
    update_3(data_3)

def fillout_3(e):
    entry_3.delete(0, END)
    entry_3.insert(0, my_list_3.get(ACTIVE))

entry_3 = Entry(interface, width=50)
entry_3.grid(row=3, column=5, padx=5, pady=5)

my_list_3: Listbox = Listbox(interface, height=20, width=50)
my_list_3.grid(row=4, column=5, padx=5, pady=5)

my_list_3.bind("<<ListboxSelect>>", fillout_3)
entry_3.bind('<KeyRelease>', check_3)

update(alist)

# OPTIE 3: op basis van fotograaf/handtekeningen/namen/personen
# OPTIE 4: overige (zoals kartelrand, KIK-nummer bv. B0...)
# KIK-nummer = "opgelet! Je geeft aan dat de foto een KIK-nummer bevat. Het originele beeld is dan ook terug te vinden in de dataabase van het KIK. Kijk daar of de foto gedateerd is."
# Zowel één datering als overzicht + waarschuwing dat je altijd moet controleren

interface.mainloop()
